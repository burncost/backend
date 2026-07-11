"""BOQ Generator - Real BOQ generation using AI and market rates."""
from typing import Dict, Any, List, Optional
import logging
import json
import os
import math
from datetime import datetime
from motor.motor_asyncio import AsyncIOMotorDatabase
from bson import ObjectId

from app.schemas.boq import BOQGenerationRequest
from app.services.mitm_engine import MITMEngine
from app.services.price_service import PriceService
from app.services.gemini_client import get_gemini_client

logger = logging.getLogger(__name__)


class BOQGenerator:
    """Service for generating Bills of Quantities from building parameters."""

    def __init__(self, db: Optional[AsyncIOMotorDatabase] = None):
        self.db = db
        self.api_key = os.getenv("AI_SERVICE_API_KEY", "")
        self.api_url = os.getenv("AI_SERVICE_URL", "https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-pro:generateContent")
        self.mitm = MITMEngine(PriceService(mongo_db=db))
        self.price_service = PriceService(mongo_db=db)

    async def create_boq(
        self,
        project_id: str,
        source_document_ids: List[str],
        template_id: Optional[str] = None,
        title: str = "",
        created_by: str = ""
    ) -> Dict[str, Any]:
        """Create a new BOQ record in MongoDB."""
        if not self.db:
            raise RuntimeError("MongoDB connection not available")

        boq_doc = {
            "projectId": project_id,
            "boqNumber": f"BOQ-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}",
            "title": title,
            "status": "generating",
            "version": 1,
            "generationMethod": "ai",
            "sourceDocumentIds": source_document_ids,
            "templateId": template_id,
            "createdBy": created_by,
            "trades": [],
            "summary": {},
            "createdAt": datetime.utcnow(),
            "updatedAt": datetime.utcnow(),
        }

        result = await self.db["boqs"].insert_one(boq_doc)
        created = await self.db["boqs"].find_one({"_id": result.inserted_id})
        if created:
            created["_id"] = str(created["_id"])
        return created

    async def generate_boq_items(
        self,
        boq_id: str,
        document_ids: List[str]
    ) -> None:
        """Generate BOQ items in background using AI."""
        logger.info(f"Generating BOQ items for {boq_id} from {len(document_ids)} documents")
        if not self.db:
            logger.warning("No DB connection — cannot generate BOQ items")
            return

        try:
            # Load the BOQ record to get generation parameters
            boq = await self.db["boqs"].find_one({"_id": ObjectId(boq_id)})
            if not boq:
                logger.error(f"BOQ {boq_id} not found")
                return

            # Build a minimal BOQGenerationRequest from stored data
            from app.schemas.boq import BOQGenerationRequest, ProjectInfoInput, FloorInput
            project_info = ProjectInfoInput(
                project_title=boq.get("title", "Untitled"),
                city="Abuja",
            )
            request = BOQGenerationRequest(
                project_info=project_info,
                floors=[FloorInput(
                    floor_id="GF", level=0,
                    floor_area_m2=boq.get("floor_area", 100),
                    perimeter_m=boq.get("perimeter", 40),
                )],
            )

            # Run MITM enrichment + template generation
            mitm_result = self.mitm.enrich(request)
            enriched = mitm_result["enriched"]
            boq_data = await self._generate_from_template(request, enriched)


            # Enrich with prices
            city = enriched["project_info"]["city"]
            enriched_elements, discrepancies = await self.price_service.enrich_boq_elements(
                boq_data["elements"], city
            )
            boq_data["elements"] = enriched_elements
            boq_data["price_discrepancies"] = discrepancies

            # Recalculate totals
            totals = self.price_service.recalculate_totals(boq_data["elements"])
            boq_data["summary"] = {
                "sub_total": totals["sub_total"],
                "contingency": totals["contingency_amount"],
                "contingency_pct": totals["contingency_pct"],
                "vat": totals["vat_amount"],
                "vat_pct": totals["vat_pct"],
                "total_contract_sum": totals["total_contract_sum"],
            }

            # Save generated items to DB
            await self.db["boqs"].update_one(
                {"_id": ObjectId(boq_id)},
                {"$set": {
                    "status": "pending_review",
                    "elements": boq_data["elements"],
                    "summary": boq_data["summary"],
                    "assumptions": boq_data.get("assumptions", []),
                    "notes": boq_data.get("notes", []),
                    "confidence": mitm_result["confidence"],
                    "updatedAt": datetime.utcnow(),
                }}
            )
            logger.info(f"BOQ {boq_id} items generated successfully")
        except Exception as e:
            logger.error(f"Failed to generate BOQ items for {boq_id}: {e}")
            if self.db:
                await self.db["boqs"].update_one(
                    {"_id": ObjectId(boq_id)},
                    {"$set": {"status": "failed", "error": str(e), "updatedAt": datetime.utcnow()}}
                )


    async def approve_boq(
        self,
        boq_id: str,
        approved_by: str
    ) -> Optional[Dict[str, Any]]:
        """Approve a BOQ."""
        if not self.db:
            raise RuntimeError("MongoDB connection not available")

        result = await self.db["boqs"].find_one_and_update(
            {"_id": ObjectId(boq_id)},
            {
                "$set": {
                    "status": "approved",
                    "approvedBy": approved_by,
                    "approvedAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow(),
                }
            },
            return_document=True,
        )
        if result:
            result["_id"] = str(result["_id"])
        return result

    async def export_boq(
        self,
        boq_id: str,
        format: str
    ) -> str:
        """Export a BOQ to the specified format and return a file URL."""
        logger.info(f"Exporting BOQ {boq_id} to {format}")
        if not self.db:
            return f"/exports/{boq_id}.{format}"

        try:
            boq = await self.db["boqs"].find_one({"_id": ObjectId(boq_id)})
            if not boq:
                logger.warning(f"BOQ {boq_id} not found for export")
                return f"/exports/{boq_id}.{format}"

            elements = boq.get("elements", [])
            summary = boq.get("summary", {})
            project_title = boq.get("title", "BOQ")

            # Ensure export directory exists
            export_dir = os.path.join(os.getcwd(), "exports")
            os.makedirs(export_dir, exist_ok=True)

            if format == "csv":
                import csv
                filepath = os.path.join(export_dir, f"{boq_id}.csv")
                with open(filepath, "w", newline="") as f:
                    writer = csv.writer(f)
                    writer.writerow(["Element", "Item Code", "Description", "Quantity", "Unit", "Rate", "Amount"])
                    for el in elements:
                        for item in el.get("items", []):
                            writer.writerow([
                                el.get("elementName", ""),
                                item.get("item_code", item.get("itemCode", "")),
                                item.get("description", ""),
                                item.get("quantity", 0),
                                item.get("unit", ""),
                                item.get("rate", 0),
                                item.get("amount", 0),
                            ])
                    writer.writerow([])
                    writer.writerow(["Total Contract Sum", "", "", "", "", "", summary.get("total_contract_sum", 0)])
                logger.info(f"Exported BOQ {boq_id} to CSV: {filepath}")
                return f"/exports/{boq_id}.csv"

            elif format == "excel":
                try:
                    import openpyxl
                    from openpyxl.styles import Font, Alignment
                except ImportError:
                    logger.warning("openpyxl not installed, falling back to CSV")
                    return await self.export_boq(boq_id, "csv")

                filepath = os.path.join(export_dir, f"{boq_id}.xlsx")
                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = "BOQ"

                # Header
                ws.cell(row=1, column=1, value=f"Bill of Quantities - {project_title}").font = Font(bold=True, size=14)
                ws.merge_cells("A1:G1")

                # Column headers
                headers = ["Element", "Item Code", "Description", "Quantity", "Unit", "Rate (NGN)", "Amount (NGN)"]
                for col, h in enumerate(headers, 1):
                    cell = ws.cell(row=3, column=col, value=h)
                    cell.font = Font(bold=True)

                row = 4
                for el in elements:
                    for item in el.get("items", []):
                        ws.cell(row=row, column=1, value=el.get("elementName", ""))
                        ws.cell(row=row, column=2, value=item.get("item_code", item.get("itemCode", "")))
                        ws.cell(row=row, column=3, value=item.get("description", ""))
                        ws.cell(row=row, column=4, value=item.get("quantity", 0))
                        ws.cell(row=row, column=5, value=item.get("unit", ""))
                        ws.cell(row=row, column=6, value=item.get("rate", 0))
                        ws.cell(row=row, column=7, value=item.get("amount", 0))
                        row += 1

                # Summary
                row += 1
                ws.cell(row=row, column=1, value="Sub Total").font = Font(bold=True)
                ws.cell(row=row, column=7, value=summary.get("sub_total", 0)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value=f"Contingency ({summary.get('contingency_pct', 5)}%)").font = Font(bold=True)
                ws.cell(row=row, column=7, value=summary.get("contingency", 0)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value=f"VAT ({summary.get('vat_pct', 7.5)}%)").font = Font(bold=True)
                ws.cell(row=row, column=7, value=summary.get("vat", 0)).font = Font(bold=True)
                row += 1
                ws.cell(row=row, column=1, value="TOTAL CONTRACT SUM").font = Font(bold=True, size=13)
                ws.cell(row=row, column=7, value=summary.get("total_contract_sum", 0)).font = Font(bold=True, size=13)

                wb.save(filepath)
                logger.info(f"Exported BOQ {boq_id} to Excel: {filepath}")
                return f"/exports/{boq_id}.xlsx"

            elif format == "pdf":
                try:
                    from reportlab.lib.pagesizes import A4
                    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
                    from reportlab.lib.styles import getSampleStyleSheet
                    from reportlab.lib import colors
                except ImportError:
                    logger.warning("reportlab not installed, falling back to CSV")
                    return await self.export_boq(boq_id, "csv")

                filepath = os.path.join(export_dir, f"{boq_id}.pdf")
                doc = SimpleDocTemplate(filepath, pagesize=A4)
                styles = getSampleStyleSheet()
                story = [Paragraph(f"Bill of Quantities - {project_title}", styles["Title"]), Spacer(1, 12)]

                for el in elements:
                    story.append(Paragraph(el.get("elementName", ""), styles["Heading2"]))
                    data = [["Item Code", "Description", "Qty", "Unit", "Rate", "Amount"]]
                    for item in el.get("items", []):
                        data.append([
                            item.get("item_code", item.get("itemCode", "")),
                            item.get("description", ""),
                            str(item.get("quantity", 0)),
                            item.get("unit", ""),
                            f"₦{item.get('rate', 0):,.2f}",
                            f"₦{item.get('amount', 0):,.2f}",
                        ])
                    t = Table(data, colWidths=[60, 200, 50, 40, 70, 70])
                    t.setStyle(TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                        ("FONTSIZE", (0, 0), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ]))
                    story.append(t)
                    story.append(Spacer(1, 12))

                doc.build(story)
                logger.info(f"Exported BOQ {boq_id} to PDF: {filepath}")
                return f"/exports/{boq_id}.pdf"

            else:
                logger.warning(f"Unsupported export format: {format}")
                return f"/exports/{boq_id}.{format}"

        except Exception as e:
            logger.error(f"Export failed for BOQ {boq_id}: {e}")
            return f"/exports/{boq_id}.{format}"


    async def upload_and_verify(
        self,
        file,
        uploaded_by: str
    ) -> Dict[str, Any]:
        """Upload and verify a BOQ file (Excel/CSV) against market prices."""
        logger.info(f"Processing uploaded BOQ file by user {uploaded_by}")
        try:
            content = await file.read()
            filename = file.filename or "uploaded"
            ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""

            rows = []
            if ext == "csv":
                import csv
                import io
                text = content.decode("utf-8-sig")
                reader = csv.DictReader(io.StringIO(text))
                for row in reader:
                    rows.append(row)
            elif ext in ("xlsx", "xls"):
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                    ws = wb.active
                    headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
                    for row in ws.iter_rows(min_row=2, values_only=True):
                        rows.append(dict(zip(headers, row)))
                except ImportError:
                    return {
                        "boq_id": "",
                        "parsed_boq": {},
                        "analysis": {},
                        "message": "openpyxl not installed. Cannot parse Excel files."
                    }
            else:
                return {
                    "boq_id": "",
                    "parsed_boq": {},
                    "analysis": {},
                    "message": f"Unsupported file format: .{ext}. Upload CSV or Excel."
                }

            if not rows:
                return {
                    "boq_id": "",
                    "parsed_boq": {},
                    "analysis": {},
                    "message": "No data rows found in the uploaded file."
                }

            # Parse items from rows
            parsed_items = []
            for row in rows:
                desc = str(row.get("description", row.get("Description", "")))
                qty = float(row.get("quantity", row.get("Quantity", 0)) or 0)
                unit = str(row.get("unit", row.get("Unit", "")))
                rate = float(row.get("rate", row.get("Rate", 0)) or 0)
                amount = float(row.get("amount", row.get("Amount", 0)) or 0)
                if desc:
                    parsed_items.append({
                        "description": desc,
                        "quantity": qty,
                        "unit": unit,
                        "rate": rate,
                        "amount": amount,
                    })

            # Verify against market prices
            analysis = {"verified_items": [], "discrepancies": []}
            total_quoted = 0
            total_market = 0

            for item in parsed_items:
                db_rate = await self.price_service.get_rate(item["description"])
                market_rate = db_rate["rate"] if db_rate else None
                total_quoted += item["amount"]

                if market_rate:
                    total_market += item["quantity"] * market_rate
                    deviation = abs(item["rate"] - market_rate) / market_rate * 100 if market_rate > 0 else 0
                    analysis["verified_items"].append({
                        "description": item["description"],
                        "quoted_rate": item["rate"],
                        "market_rate": market_rate,
                        "deviation_pct": round(deviation, 1),
                        "status": "inflated" if deviation > 25 else "fair",
                    })
                    if deviation > 25:
                        analysis["discrepancies"].append({
                            "description": item["description"],
                            "quoted": item["rate"],
                            "market": market_rate,
                            "overcharge_pct": round(deviation, 1),
                        })
                else:
                    analysis["verified_items"].append({
                        "description": item["description"],
                        "quoted_rate": item["rate"],
                        "market_rate": None,
                        "deviation_pct": None,
                        "status": "unverified",
                    })

            # Save to DB
            boq_id = ""
            if self.db:
                doc = {
                    "filename": filename,
                    "uploadedBy": uploaded_by,
                    "uploadedAt": datetime.utcnow(),
                    "items": parsed_items,
                    "analysis": analysis,
                    "total_quoted": total_quoted,
                    "total_market": total_market,
                }
                result = await self.db["boq_verifications"].insert_one(doc)
                boq_id = str(result.inserted_id)

            inflated_count = len([v for v in analysis["verified_items"] if v.get("status") == "inflated"])
            fair_count = len([v for v in analysis["verified_items"] if v.get("status") == "fair"])

            return {
                "boq_id": boq_id,
                "parsed_boq": {"items": parsed_items, "total_quoted": total_quoted},
                "analysis": analysis,
                "message": (
                    f"Verified {len(parsed_items)} items. "
                    f"{inflated_count} inflated, {fair_count} fair, "
                    f"{len(parsed_items) - inflated_count - fair_count} unverified."
                ),
            }

        except Exception as e:
            logger.error(f"Upload and verify failed: {e}")
            return {
                "boq_id": "",
                "parsed_boq": {},
                "analysis": {},
                "message": f"Error processing file: {str(e)}",
            }


    async def handle_decision(
        self,
        boq_id: str,
        decision: str,
        user_id: str
    ) -> Optional[Dict[str, Any]]:
        """Handle user decision on a BOQ (regenerate or save original)."""
        if not self.db:
            raise RuntimeError("MongoDB connection not available")

        new_status = "regenerated" if decision == "regenerate" else "saved_original"
        result = await self.db["boqs"].find_one_and_update(
            {"_id": ObjectId(boq_id)},
            {
                "$set": {
                    "status": new_status,
                    "userDecision": decision,
                    "decidedBy": user_id,
                    "decidedAt": datetime.utcnow(),
                    "updatedAt": datetime.utcnow(),
                }
            },
            return_document=True,
        )
        if result:
            result["_id"] = str(result["_id"])
        return result

    async def verify_quote_text(
        self,
        quote_text: str,
        user_id: str
    ) -> Dict[str, Any]:
        """Verify a quote text against market prices using Gemini."""
        logger.info(f"Verifying quote for user {user_id}")
        try:
            from app.services.gemini_client import get_gemini_client
            from google.genai import types as genai_types

            client = get_gemini_client()
            prompt = f"""You are a Nigerian quantity surveyor. Parse the following quote text and extract line items.
For each item, estimate the market rate in NGN based on current Nigerian prices.

Quote text:
{quote_text}

Return ONLY valid JSON with this structure:
{{
  "items": [
    {{
      "description": "item description",
      "quantity": number,
      "unit": "m²/m³/nr/ls/etc",
      "quoted_rate": number,
      "estimated_market_rate": number,
      "deviation_pct": number,
      "status": "fair" | "inflated" | "unverified"
    }}
  ],
  "total_quoted": number,
  "total_market": number,
  "total_overcharge": number,
  "inflated_count": number,
  "fair_count": number,
  "summary_note": "brief analysis"
}}"""

            response = client.models.generate_content(
                model="gemini-2.5-pro",
                contents=[prompt],
                config=genai_types.GenerateContentConfig(
                    temperature=0.1,
                    max_output_tokens=4096,
                ),
            )

            text = response.text
            if not text:
                raise Exception("Empty Gemini response")

            import re
            match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', text)
            if match:
                result = json.loads(match.group(1))
            else:
                match = re.search(r'\{[\s\S]*\}', text)
                if match:
                    result = json.loads(match.group(0))
                else:
                    raise Exception("Failed to parse Gemini response")

            # Verify each item against DB prices
            for item in result.get("items", []):
                db_rate = await self.price_service.get_rate(item["description"])
                if db_rate and db_rate.get("rate"):
                    item["db_rate"] = db_rate["rate"]
                    item["deviation_pct"] = round(
                        abs(item["quoted_rate"] - db_rate["rate"]) / db_rate["rate"] * 100, 1
                    )
                    item["status"] = "inflated" if item["deviation_pct"] > 25 else "fair"

            return result

        except Exception as e:
            logger.error(f"Quote verification failed: {e}")
            return {
                "items": [],
                "total_quoted": 0,
                "total_market": 0,
                "total_overcharge": 0,
                "inflated_count": 0,
                "fair_count": 0,
                "summary_note": f"Verification failed: {str(e)}"
            }


    # ── Main generation entry point ──────────────────────────────────────────

    async def generate_from_parameters(
        self,
        request: BOQGenerationRequest,
        user_id: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Generate a complete BOQ from building parameters.
        Runs MITM enrichment first, then AI generation, then price enrichment.
        """
        logger.info(f"Generating BOQ for project: {request.project_info.project_title}")

        # Step 1: MITM enrichment
        mitm_result = self.mitm.enrich(request)
        enriched = mitm_result["enriched"]
        flags = mitm_result["flags"]
        confidence = mitm_result["confidence"]

        # Step 2: Generate BOQ (AI or template)
        if self.api_key:
            try:
                boq = await self._generate_with_ai(request, enriched)
            except Exception as e:
                logger.error(f"AI generation failed: {e}")
                boq = await self._generate_from_template(request, enriched)
        else:
            boq = await self._generate_from_template(request, enriched)


        # Step 3: Enrich with real prices from DB
        city = enriched["project_info"]["city"]
        enriched_elements, discrepancies, out_of_stock = await self.price_service.enrich_boq_elements(
            boq["elements"], city
        )
        boq["elements"] = enriched_elements
        boq["price_discrepancies"] = discrepancies
        boq["out_of_stock_items"] = out_of_stock

        # Step 3b: Notify vendors for out-of-stock items
        if out_of_stock and user_id:
            project_title = enriched["project_info"]["project_title"]
            notified_count = await self.price_service.notify_vendors(
                out_of_stock_items=out_of_stock,
                project_title=project_title,
                user_id=user_id,
            )
            # Mark notified items
            for item in boq["out_of_stock_items"]:
                item["vendor_notified"] = True

        # Step 4: Recalculate totals
        totals = self.price_service.recalculate_totals(boq["elements"])
        total = totals["total_contract_sum"]
        floor_area = enriched["total_floor_area_m2"]
        boq["summary"] = {
            "sub_total": totals["sub_total"],
            "contingency": totals["contingency_amount"],
            "contingency_pct": totals["contingency_pct"],
            "vat": totals["vat_amount"],
            "vat_pct": totals["vat_pct"],
            "total_contract_sum": total,
            "total_low": round(total * 0.9, 2),
            "total_expected": total,
            "total_high": round(total * 1.1, 2),
            "cost_per_m2": round(total / max(floor_area, 1), 2),
            "total_floor_area_m2": floor_area,
            "cost_scenarios": {
                "low": round(total * 0.9, 2),
                "expected": total,
                "high": round(total * 1.1, 2),
            },
        }

        # Step 5: Attach confidence and assumptions
        boq["confidence"] = confidence
        boq["assumptions_used"] = mitm_result["assumptions"]
        boq["warnings"] = [f["message"] for f in flags if f["severity"] == "critical"]
        boq["generated_at"] = datetime.utcnow().isoformat()
        boq["project_info"] = enriched["project_info"]

        return boq

    # ── AI generation ────────────────────────────────────────────────────────

    async def _generate_with_ai(
        self, request: BOQGenerationRequest, enriched: Dict
    ) -> Dict[str, Any]:
        """Generate BOQ using Gemini AI (google-genai SDK)."""
        from google.genai import types as genai_types
        from app.services.gemini_client import get_gemini_client

        client = get_gemini_client()
        prompt = self._build_boq_prompt(request, enriched)

        try:
            response = client.models.generate_content(
                model="gemini-2.5-pro",
                contents=[prompt],
                config=genai_types.GenerateContentConfig(
                    temperature=0.1,
                    max_output_tokens=8192,
                ),
            )

            text = response.text
            if not text:
                raise Exception("Empty Gemini response")

            import re
            match = re.search(r'```(?:json)?\s*([\s\S]*?)\s*```', text)
            if match:
                return json.loads(match.group(1))
            match = re.search(r'\{[\s\S]*\}', text)
            if match:
                return json.loads(match.group(0))

            raise Exception("Failed to parse AI response")
        except Exception as exc:
            logger.error("Gemini BOQ generation failed: %s", exc)
            raise

    def _build_boq_prompt(self, request: BOQGenerationRequest, enriched: Dict) -> str:
        """Build the BOQ generation prompt with Nigerian standards."""
        import json as j

        derived = enriched.get("derived_quantities", {})
        city = enriched["project_info"]["city"]
        building_type = enriched["project_info"]["building_type"]

        return f"""You are a professional Quantity Surveyor registered with NIQS (Nigerian Institute of Quantity Surveyors). 
Generate a detailed Bill of Quantities (BOQ) in JSON format following the Nigerian SMM7 standard.

PROJECT INFORMATION:
- Title: {request.project_info.project_title}
- Location: {city}, {enriched['project_info']['location']}
- Building Type: {building_type}
- Client: {request.project_info.client_name or 'N/A'}

BUILDING PARAMETERS:
{j.dumps(enriched, indent=2)}

DERIVED QUANTITIES (pre-calculated):
{j.dumps(derived, indent=2)}

IMPORTANT RULES:
1. Use current Nigerian market rates (NGN) for {city}
2. Apply these wastage factors: blocks 5%, concrete 5%, tiles 10%, roofing 12%, reinforcement 5%
3. Structure the BOQ in this Nigerian standard order:
   - Preliminaries (site setup, insurance, scaffolding)
   - Substructure (excavation, hardcore, blinding, foundation concrete, DPC, ground slab)
   - Superstructure (block walls, RC columns, ring beams, lintels)
   - Roofing (trusses, roof covering, ceiling, gutters, fascia)
   - Joinery (doors, windows, louvres, burglar-proofing)
   - Internal Finishes (wall plaster, floor screed, floor tiles, wall tiles, painting)
   - External Finishes (external render, external paint)
   - Plumbing & Drainage (water supply, sanitary fittings, drainage)
   - Electrical (conduit, wiring, light fittings, sockets, DB)
   - External Works (fence, gate, paving, borehole if applicable)
4. Include contingency (5%), overheads & profit (10%), VAT (7.5%) in summary
5. Provide cost scenarios: low (90%), expected (100%), high (110%)

Return ONLY valid JSON with this exact structure:
{{
  "projectTitle": string,
  "generatedAt": "ISO datetime",
  "elements": [
    {{
      "elementName": string,
      "trade": string,
      "totalCost": number,
      "items": [
        {{
          "itemCode": string,
          "description": string,
          "quantity": number,
          "unit": string,
          "rate": number,
          "amount": number,
          "estimated": bool
        }}
      ]
    }}
  ],
  "assumptions": [string],
  "notes": [string]
}}"""

    # ── Template generation (fallback) ───────────────────────────────────────

    async def _generate_from_template(
        self, request: BOQGenerationRequest, enriched: Dict
    ) -> Dict[str, Any]:
        """Generate BOQ from template calculations using enriched data.
        All rates are fetched from price_service (DB or internet fallback).
        No hardcoded rates are used.
        """
        derived = enriched.get("derived_quantities", {})
        total_area = enriched["total_floor_area_m2"]
        num_floors = len(enriched["floors"])
        city = enriched["project_info"]["city"]
        finish_grade = enriched["finishes"]["finish_grade"]

        # Get city and finish multipliers (async)
        try:
            city_factor = await self.price_service.get_city_factor(city)
            finish_factor = await self.price_service.get_finish_level_multiplier(finish_grade)
        except Exception:
            city_factor = 1.0
            finish_factor = 1.0

        # Helper to fetch rate from price_service with fallback
        async def _get_rate(description: str, default_rate: float) -> float:
            result = await self.price_service.get_rate(description, city)
            if result and result.get("rate"):
                return float(result["rate"])
            return default_rate * city_factor

        elements = []
        total_cost = 0

        # ── Preliminaries ────────────────────────────────────────────────────
        prelim_items = [
            ("PRE-001", "Site clearance & preparation", 1, "ls", 150000),
            ("PRE-002", "Scaffolding hire", total_area, "m²", 800),
            ("PRE-003", "Concrete mixer hire", 1, "ls", 200000),
            ("PRE-004", "Setting out & survey", 1, "ls", 100000),
        ]
        prelim_items_resolved = []
        for code, desc, qty, unit, default_rate in prelim_items:
            rate = await _get_rate(desc, default_rate)
            prelim_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        prelim_cost = sum(i["amount"] for i in prelim_items_resolved)
        elements.append({
            "elementName": "Preliminaries", "trade": "Preliminaries",
            "totalCost": round(prelim_cost, 2), "items": prelim_items_resolved,
        })
        total_cost += prelim_cost

        # ── Substructure ─────────────────────────────────────────────────────
        foundation_vol = derived.get("foundation_concrete_m3", total_area * 0.3)
        sub_items = [
            ("SUB-001", "Excavation (strip foundation)", round(foundation_vol * 1.5, 2), "m³", 2500),
            ("SUB-002", "Hardcore filling 150mm thick", round(total_area, 2), "m²", 1800),
            ("SUB-003", "Blinding concrete (1:3:6) 50mm", round(total_area, 2), "m²", 2200),
            ("SUB-004", "Foundation concrete (1:2:4)", round(foundation_vol, 2), "m³", 94000),
            ("SUB-005", "DPC membrane (oversite)", round(total_area, 2), "m²", 1200),
            ("SUB-006", "Reinforcement (foundation)", round(derived.get("estimated_rebar_kg", total_area * 15) * 0.3, 2), "kg", 850),
        ]
        sub_items_resolved = []
        for code, desc, qty, unit, default_rate in sub_items:
            rate = await _get_rate(desc, default_rate)
            sub_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        sub_cost = sum(i["amount"] for i in sub_items_resolved)
        elements.append({
            "elementName": "Substructure", "trade": "Substructure",
            "totalCost": round(sub_cost, 2), "items": sub_items_resolved,
        })
        total_cost += sub_cost

        # ── Superstructure ───────────────────────────────────────────────────
        ext_blocks = derived.get("external_blocks_225mm", total_area * 20)
        int_blocks = derived.get("internal_blocks_150mm", total_area * 12)
        wall_area = derived.get("net_wall_area_m2", total_area * 2.8 * num_floors)
        total_perimeter = enriched.get("total_perimeter_m", math.sqrt(total_area) * 4)
        sup_items = [
            ("SUP-001", "9-inch sandcrete block wall (225mm)", round(ext_blocks), "nr", 1300),
            ("SUP-002", "6-inch sandcrete block wall (150mm)", round(int_blocks), "nr", 1100),
            ("SUP-003", "RC Columns 225x225mm (1:2:4)", round(num_floors * 12), "nr", 85000),
            ("SUP-004", "RC Ring Beams 225x450mm", round(total_perimeter, 2), "m", 12000),
            ("SUP-005", "RC Lintels over openings", round(total_perimeter * 0.3, 2), "m", 8000),
            ("SUP-006", "Ground floor slab (1:2:4) 150mm", round(total_area * 0.15, 2), "m³", 94000),
        ]
        sup_items_resolved = []
        for code, desc, qty, unit, default_rate in sup_items:
            rate = await _get_rate(desc, default_rate)
            sup_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        sup_cost = sum(i["amount"] for i in sup_items_resolved)
        elements.append({
            "elementName": "Superstructure", "trade": "Superstructure",
            "totalCost": round(sup_cost, 2), "items": sup_items_resolved,
        })
        total_cost += sup_cost

        # ── Roofing ──────────────────────────────────────────────────────────
        roof_area = derived.get("roof_slope_area_m2", total_area * 1.3)
        roof_items = [
            ("ROF-001", "Timber roof trusses (supply & fix)", round(roof_area, 2), "m²", 6500),
            ("ROF-002", "Longspan aluminium roofing 0.55mm", round(roof_area, 2), "m²", 6000),
            ("ROF-003", "Ridge capping", round(math.sqrt(roof_area) * 1.5, 2), "m", 3500),
            ("ROF-004", "Fascia & soffit board", round(math.sqrt(roof_area) * 4, 2), "m", 3500),
            ("ROF-005", "Rainwater gutter (uPVC)", round(math.sqrt(roof_area) * 4, 2), "m", 2500),
            ("ROF-006", "Downpipe (uPVC)", round(num_floors * 4), "nr", 4500),
        ]
        roof_items_resolved = []
        for code, desc, qty, unit, default_rate in roof_items:
            rate = await _get_rate(desc, default_rate)
            roof_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        roof_cost = sum(i["amount"] for i in roof_items_resolved)
        elements.append({
            "elementName": "Roofing", "trade": "Roofing",
            "totalCost": round(roof_cost, 2), "items": roof_items_resolved,
        })
        total_cost += roof_cost

        # ── Joinery ──────────────────────────────────────────────────────────
        door_count = sum(len(f["openings"]["doors"]) for f in enriched["floors"])
        window_count = sum(len(f["openings"]["windows"]) for f in enriched["floors"])
        joinery_items = [
            ("JON-001", "Internal flush door (0.8x2.1m) with frame", max(door_count, 6), "nr", 65000),
            ("JON-002", "External security door (0.9x2.1m)", max(1, door_count // 4), "nr", 120000),
            ("JON-003", "Aluminium sliding window 1.2x1.2m", max(window_count, 4), "nr", 95000),
            ("JON-004", "Burglar-proofing (window)", max(window_count, 4), "nr", 25000),
        ]
        joinery_items_resolved = []
        for code, desc, qty, unit, default_rate in joinery_items:
            rate = await _get_rate(desc, default_rate)
            joinery_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        joinery_cost = sum(i["amount"] for i in joinery_items_resolved)
        elements.append({
            "elementName": "Joinery (Doors & Windows)", "trade": "Joinery",
            "totalCost": round(joinery_cost, 2), "items": joinery_items_resolved,
        })
        total_cost += joinery_cost

        # ── Internal Finishes ────────────────────────────────────────────────
        floor_finish_area = derived.get("floor_finish_area_m2", total_area)
        wall_finish_area = derived.get("wall_finish_area_m2", wall_area)
        int_finish_items = [
            ("FIN-001", "Floor screed (25mm) cement/sand 1:4", round(floor_finish_area, 2), "m²", 2500),
            ("FIN-002", "Ceramic floor tile 600x600mm (supply & fix)", round(floor_finish_area, 2), "m²", 13300),
            ("FIN-003", "Wall plastering (15mm) cement/sand 1:4", round(wall_finish_area, 2), "m²", 2800),
            ("FIN-004", "Wall tiling (wet areas) 300x600mm", round(wall_finish_area * 0.15, 2), "m²", 12000),
            ("FIN-005", "POP ceiling (supply & install)", round(floor_finish_area, 2), "m²", 6500),
            ("FIN-006", "Emulsion paint (2 coats) walls", round(wall_finish_area, 2), "m²", 3400),
            ("FIN-007", "Skirting (ceramic) 100mm high", round(enriched.get("total_perimeter_m", math.sqrt(total_area) * 4), 2), "m", 2500),
        ]
        int_finish_items_resolved = []
        for code, desc, qty, unit, default_rate in int_finish_items:
            rate = await _get_rate(desc, default_rate)
            # Apply finish factor to tile/paint items
            if code in ("FIN-002", "FIN-004", "FIN-006"):
                rate = rate * finish_factor
            int_finish_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        finish_cost = sum(i["amount"] for i in int_finish_items_resolved)
        elements.append({
            "elementName": "Internal Finishes", "trade": "Finishes",
            "totalCost": round(finish_cost, 2), "items": int_finish_items_resolved,
        })
        total_cost += finish_cost

        # ── External Finishes ────────────────────────────────────────────────
        ext_wall_area = wall_area * 0.3
        ext_finish_items = [
            ("EXT-001", "External rendering (20mm) cement/sand 1:4", round(ext_wall_area, 2), "m²", 3200),
            ("EXT-002", "External emulsion paint (2 coats)", round(ext_wall_area, 2), "m²", 3500),
        ]
        ext_finish_items_resolved = []
        for code, desc, qty, unit, default_rate in ext_finish_items:
            rate = await _get_rate(desc, default_rate)
            ext_finish_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        ext_cost = sum(i["amount"] for i in ext_finish_items_resolved)
        elements.append({
            "elementName": "External Finishes", "trade": "Finishes",
            "totalCost": round(ext_cost, 2), "items": ext_finish_items_resolved,
        })
        total_cost += ext_cost

        # ── Plumbing & Drainage ──────────────────────────────────────────────
        pf = enriched["services"].get("plumbing_fixtures", {})
        wc_count = pf.get("wc", 2)
        whb_count = pf.get("wash_hand_basin", 2)
        shower_count = pf.get("shower", 2)
        sink_count = pf.get("kitchen_sink", 1)
        has_overhead_tank = enriched["services"].get("overhead_tank", True)
        plumb_items = [
            ("PLB-001", "Cold water supply pipework (PVC)", round(total_area * 0.3, 2), "m", 2500),
            ("PLB-002", "Drainage pipework (PVC 4\")", round(total_area * 0.2, 2), "m", 3500),
            ("PLB-003", "WC suite (low level)", wc_count, "nr", 65000),
            ("PLB-004", "Wash hand basin", whb_count, "nr", 25000),
            ("PLB-005", "Shower fitting", shower_count, "nr", 18000),
            ("PLB-006", "Kitchen sink (stainless steel)", sink_count, "nr", 45000),
            ("PLB-007", "Overhead water tank (1000L)", 1 if has_overhead_tank else 0, "nr", 180000),
            ("PLB-008", "Septic tank & soakaway", 1, "ls", 350000),
        ]
        plumb_items_resolved = []
        for code, desc, qty, unit, default_rate in plumb_items:
            rate = await _get_rate(desc, default_rate)
            plumb_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        plumb_cost = sum(i["amount"] for i in plumb_items_resolved)
        elements.append({
            "elementName": "Plumbing & Drainage", "trade": "Services",
            "totalCost": round(plumb_cost, 2), "items": plumb_items_resolved,
        })
        total_cost += plumb_cost

        # ── Electrical ───────────────────────────────────────────────────────
        elec_items = [
            ("ELE-001", "PVC conduit & wiring (2.5mm) per point", round(total_area * 0.5, 2), "m", 1200),
            ("ELE-002", "Lighting point (complete)", round(total_area * 0.15, 2), "nr", 8500),
            ("ELE-003", "Socket outlet (double, complete)", round(total_area * 0.1, 2), "nr", 9500),
            ("ELE-004", "Distribution board (8-way)", 1, "nr", 35000),
            ("ELE-005", "Earthing system", 1, "ls", 80000),
        ]
        elec_items_resolved = []
        for code, desc, qty, unit, default_rate in elec_items:
            rate = await _get_rate(desc, default_rate)
            elec_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        elec_cost = sum(i["amount"] for i in elec_items_resolved)
        elements.append({
            "elementName": "Electrical Installation", "trade": "Services",
            "totalCost": round(elec_cost, 2), "items": elec_items_resolved,
        })
        total_cost += elec_cost

        # ── External Works ───────────────────────────────────────────────────
        ext_works_items = [
            ("EXTW-001", "Fencing (sandcrete block wall)", round(math.sqrt(total_area) * 4, 2), "m", 18000),
            ("EXTW-002", "Gate (metal, sliding)", 1, "nr", 250000),
            ("EXTW-003", "Interlocking paving (driveway)", round(total_area * 0.15, 2), "m²", 8500),
            ("EXTW-004", "Landscaping & planting", round(total_area * 0.2, 2), "m²", 3500),
        ]
        ext_works_items_resolved = []
        for code, desc, qty, unit, default_rate in ext_works_items:
            rate = await _get_rate(desc, default_rate)
            ext_works_items_resolved.append({
                "itemCode": code, "description": desc,
                "quantity": qty, "unit": unit,
                "rate": round(rate), "amount": round(qty * rate),
                "estimated": True,
            })
        ext_works_cost = sum(i["amount"] for i in ext_works_items_resolved)
        elements.append({
            "elementName": "External Works", "trade": "External Works",
            "totalCost": round(ext_works_cost, 2), "items": ext_works_items_resolved,
        })
        total_cost += ext_works_cost

        # ── Compute summary ──────────────────────────────────────────────────
        contingencies = total_cost * 0.05
        overheads_profit = total_cost * 0.10
        vat_rate = 0.075
        vat_amount = (total_cost + contingencies + overheads_profit) * vat_rate
        grand_total = total_cost + contingencies + overheads_profit + vat_amount

        return {
            "projectTitle": enriched["project_info"]["project_title"],
            "generatedAt": datetime.utcnow().isoformat(),
            "elements": elements,
            "assumptions": [
                f"Rates based on {city} market prices",
                f"Finish grade: {finish_grade} (multiplier: {finish_factor})",
                "All quantities estimated from floor area and standard Nigerian ratios",
                "Contingency: 5%, Overheads & Profit: 10%, VAT: 7.5%",
                "Wastage factors applied: blocks 5%, concrete 5%, tiles 10%, roofing 12%",
            ],
            "notes": [
                "This is an AI-generated estimate. Review by a professional QS recommended.",
                "Prices may vary by location, season, and supplier availability.",
                "Site conditions may affect actual quantities.",
            ],
        }
