"""Document Service - Real document processing with CAD/PDF parsing and AI analysis."""
from typing import Dict, Any, Optional, List
import logging
import os
from datetime import datetime

from app.services.cad_parser import CADParser
from app.services.pdf_extractor import PDFExtractor
from app.services.ai_service import AIService
from app.utils.storage import StorageService

logger = logging.getLogger(__name__)


class DocumentService:
    """Service for processing uploaded construction documents."""

    def __init__(self):
        self.cad_parser = CADParser()
        self.pdf_extractor = PDFExtractor()
        self.ai_service = AIService()
        self.storage = StorageService()

    async def process_document(
        self,
        file_content: bytes,
        file_name: str,
        file_type: str,
        project_id: str,
        uploaded_by: str
    ) -> Dict[str, Any]:
        """Process an uploaded document: parse, analyze with AI, store results."""
        logger.info(f"Processing document: {file_name} (type: {file_type})")

        # Upload to cloud storage
        storage_path = f"projects/{project_id}/documents/{file_name}"
        file_url = await self.storage.upload_file(
            file=file_content,
            folder=f"projects/{project_id}",
            filename=file_name
        )

        # Parse based on file type
        extracted_metadata = {}
        if file_type in [".dwg", ".dxf", ".rvt", ".ifc"]:
            extracted_metadata = await self.cad_parser.parse(file_content, file_type)
        elif file_type == ".pdf":
            extracted_metadata = await self.pdf_extractor.extract(file_content)
        elif file_type in [".xlsx", ".xls"]:
            extracted_metadata = await self._parse_excel(file_content, file_name)

        # AI analysis
        ai_analysis = await self.ai_service.analyze_document(
            file_content=file_content,
            file_type=file_type,
            extracted_metadata=extracted_metadata
        )

        # Generate thumbnail for PDFs
        thumbnail_url = None
        if file_type == ".pdf":
            thumbnail_bytes = await self.pdf_extractor.generate_thumbnail(file_content)
            if thumbnail_bytes:
                thumbnail_url = await self.storage.upload_bytes(
                    content=thumbnail_bytes,
                    path=f"thumbnails/{project_id}/{file_name.replace('.pdf', '.png')}",
                    content_type="image/png"
                )

        return {
            "fileName": file_name,
            "fileType": file_type,
            "fileSize": len(file_content),
            "fileUrl": file_url,
            "projectId": project_id,
            "uploadedBy": uploaded_by,
            "uploadedAt": datetime.utcnow().isoformat(),
            "extractedMetadata": extracted_metadata,
            "aiAnalysis": ai_analysis,
            "thumbnailUrl": thumbnail_url,
            "status": "processed" if ai_analysis.get("processed") else "partial",
        }

    async def _parse_excel(self, content: bytes, file_name: str) -> Dict[str, Any]:
        """Parse Excel BOQ files."""
        try:
            import openpyxl
            from io import BytesIO

            wb = openpyxl.load_workbook(BytesIO(content), data_only=True)
            sheets_info = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = []
                for row in ws.iter_rows(values_only=True, max_row=50):
                    rows.append([str(c) if c is not None else "" for c in row])

                sheets_info.append({
                    "name": sheet_name,
                    "rowCount": ws.max_row,
                    "colCount": ws.max_column,
                    "preview": rows[:10] if rows else [],
                })

            return {
                "sheets": sheets_info,
                "sheetCount": len(sheets_info),
                "detectedType": "bill_of_quantities" if any(
                    any(kw in str(s["name"]).lower() for kw in ["boq", "bill", "quantity", "pricing"])
                    for s in sheets_info
                ) else "unknown",
            }

        except ImportError:
            logger.warning("openpyxl not installed, returning basic metadata")
            return {
                "sheets": [{"name": "Sheet1", "rowCount": 0, "colCount": 0, "preview": []}],
                "sheetCount": 1,
                "detectedType": "unknown",
            }
        except Exception as e:
            logger.error(f"Excel parsing failed: {str(e)}")
            return {"error": str(e), "detectedType": "unknown"}

    async def extract_boq_from_document(
        self,
        document_id: str,
        extracted_metadata: Dict[str, Any],
        ai_analysis: Dict[str, Any]
    ) -> Optional[Dict[str, Any]]:
        """Extract BOQ data from processed document analysis."""
        logger.info(f"Extracting BOQ from document: {document_id}")

        # Try to get BOQ data from AI analysis
        if ai_analysis.get("processed") and ai_analysis.get("detectedElements"):
            return self._build_boq_from_analysis(ai_analysis)

        # Try to get from extracted metadata (Excel tables)
        if extracted_metadata.get("detectedType") == "bill_of_quantities":
            return self._build_boq_from_excel(extracted_metadata)

        return None

    def _build_boq_from_analysis(self, analysis: Dict[str, Any]) -> Dict[str, Any]:
        """Build BOQ structure from AI analysis results."""
        elements = []
        total = 0

        for elem in analysis.get("detectedElements", []):
            rate = self._get_material_rate(elem.get("elementType", ""))
            amount = elem.get("totalQuantity", 0) * rate
            elements.append({
                "elementName": elem.get("elementType", "Unknown"),
                "quantity": elem.get("totalQuantity", 0),
                "unit": elem.get("unit", "nr"),
                "rate": rate,
                "amount": amount,
            })
            total += amount

        return {
            "elements": elements,
            "totalContractSum": total,
            "source": "ai_analysis",
        }

    def _build_boq_from_excel(self, metadata: Dict[str, Any]) -> Dict[str, Any]:
        """Build BOQ structure from Excel data."""
        elements = []
        total = 0

        for sheet in metadata.get("sheets", []):
            for row in sheet.get("preview", [])[1:]:  # Skip header
                if len(row) >= 4:
                    try:
                        qty = float(row[2]) if row[2] else 0
                        rate = float(row[3]) if row[3] else 0
                        amount = qty * rate
                        elements.append({
                            "elementName": str(row[0]) if row[0] else "Item",
                            "description": str(row[1]) if len(row) > 1 and row[1] else "",
                            "quantity": qty,
                            "unit": str(row[1]) if len(row) > 1 and row[1] else "nr",
                            "rate": rate,
                            "amount": amount,
                        })
                        total += amount
                    except (ValueError, IndexError):
                        continue

        return {
            "elements": elements,
            "totalContractSum": total,
            "source": "excel_extraction",
        }

    def _get_material_rate(self, element_type: str) -> float:
        """Get estimated rate for a building element type."""
        rates = {
            "external_wall": 8500,
            "internal_wall": 7200,
            "slab": 15000,
            "foundation": 45000,
            "column": 85000,
            "beam": 52000,
            "roof": 18000,
            "floor_finish": 7500,
            "wall_finish": 3500,
            "door": 45000,
            "window": 35000,
        }
        return rates.get(element_type, 10000)
